import os
import pandas as pd
import json
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Define the directory where the JSON file and images are located
data_dir = '../output/'

# Read the data from the JSON file
with open(os.path.join(data_dir, 'cell_info.json'), 'r') as f:
    data = json.load(f)

# Create a new Workbook
wb = Workbook()
ws = wb.active

# Process each cell
cells = sorted(data['cells'], key=lambda cell: (cell['y'], cell['x']))
rows = sorted(list(set(cell['y'] for cell in cells)))
cols = sorted(list(set(cell['x'] for cell in cells)))
for cell_info in cells:
    # Get cell position based on its bounding box
    row = rows.index(cell_info['y']) + 1
    col = cols.index(cell_info['x']) + 1

    # Create an Image object from the file
    img = Image(os.path.join(data_dir, cell_info['img_file']))

    # Add the image to the worksheet at the appropriate cell
    cell_address = f"{get_column_letter(col)}{row}"
    ws.add_image(img, cell_address)

# Save the Workbook to an Excel file
wb.save('output.xlsx')
